import os
import datetime
import random
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import django
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "timeforge.settings.base")
django.setup()

from django.core.exceptions import ValidationError
from django.contrib.auth import get_user_model
from core.models import Department, Room, School, Session
from academics.models import (
    Subject, Course, CourseLevel, CourseLevelOffering, TeacherProfile, ClassSession,
)
from scheduling.models import TeacherAvailability

User = get_user_model()

BASE_DIR = Path(__file__).resolve().parent
SUBJECTS_XLSX = BASE_DIR / "data" / "KU_SOE_Master_Database.xlsx"
SEMESTERS_XLSX = BASE_DIR / "data" / "semesters.xlsx"
TEACHERS_TXT = BASE_DIR / "data" / "teachers.txt"

# Spreadsheet program label → Course.code
PROGRAM_TO_COURSE_CODE = {
    "BE Computer Engineering": "BE-CE",
    "BSc Computer Science": "BSC-CS",
    "Artificial Intelligence": "BTECH-AI",
    "Geomatics Engineering": "BE-GE",
    "Chemical Engineering": "BE-CHE",
    "Civil Engineering": "BE-CIV",
    "Communication Engineering": "BE-COMM",
}

# Excel department label → seeded Department.code
EXCEL_DEPT_TO_CODE = {
    "Computer Science & Engineering": "CSE",
    "Civil Engineering": "CE",
    "Electrical & Electronics Engineering": "EEE",
    "Chemical Engineering": "CHE",
    "Geomatics Engineering": "GE",
}


def _parse_credit(value, default=Decimal("3.00")):
    if value is None or str(value).strip() in ("", "None"):
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return default


def _periods_per_week_from_credits(credit):
    """
    1 credit = 16 contact hours / semester; with ~1.5h periods that yields
    periods_per_week = credits * 2/3 (so 3 credits → 2 periods/week).
    """
    credit = _parse_credit(credit)
    periods = (credit * Decimal("2") / Decimal("3")).quantize(
        Decimal("1"), rounding=ROUND_HALF_UP
    )
    return max(1, int(periods))


def _absolute_semester(year, semester):
    """year 3, semester 1 → absolute semester 5."""
    return (int(year) - 1) * 2 + int(semester)


def _resolve_excel_departments(raw, departments_by_code):
    """Split comma-separated Excel dept labels into Department instances."""
    if not raw:
        return []
    resolved = []
    seen = set()
    for part in str(raw).split(","):
        label = part.strip()
        if not label:
            continue
        code = EXCEL_DEPT_TO_CODE.get(label)
        if code is None:
            raise ValueError(f"Unmapped Excel department label: {label!r}")
        if code in seen:
            continue
        seen.add(code)
        resolved.append(departments_by_code[code])
    return resolved


def seed_subjects_from_excel(departments_by_code):
    """Load KU SOE subjects from the scraped workbook."""
    if not SUBJECTS_XLSX.is_file():
        raise FileNotFoundError(
            f"Missing subjects workbook at {SUBJECTS_XLSX}. "
            "Place KU_SOE_Master_Database.xlsx under data/."
        )

    wb = load_workbook(SUBJECTS_XLSX, read_only=True, data_only=True)
    try:
        rows = list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()

    created = 0
    skipped = 0
    for row in rows:
        if not row or not row[0]:
            skipped += 1
            continue
        code = str(row[0]).strip()
        name = str(row[1] or "").strip()
        if not name:
            skipped += 1
            continue
        credit = _parse_credit(row[2] if len(row) > 2 else None)
        depts = _resolve_excel_departments(row[3] if len(row) > 3 else None, departments_by_code)
        if not depts:
            skipped += 1
            print(f"  SKIP {code}: no resolvable departments ({row[3]!r})")
            continue

        subject = Subject.objects.create(
            name=name[:150],
            code=code[:20],
            credit_hours=credit,
            lecture_hours_per_week=max(1, int(credit)),
            lab_hours_per_week=0,
            is_active=True,
        )
        subject.departments.set(depts)
        created += 1

    print(f"Subjects seeded from Excel: {created} created, {skipped} skipped.")
    return created


def _parse_teacher_line(line):
    """
    Return (title, first_name, last_name, is_visiting) from a roster line.

    Rules:
      Prof. Dr. … → title Prof. Dr.
      Dr. …       → title Dr.
      Mr./Ms. …   → title Lecturer
      no salutation → title blank, visiting faculty
    """
    raw = " ".join(str(line).split())
    if not raw:
        return None

    title = ""
    is_visiting = False
    name = raw

    if raw.startswith("Prof. Dr. "):
        title = TeacherProfile.Title.PROF_DR
        name = raw[len("Prof. Dr. "):]
    elif raw.startswith("Prof.Dr. "):
        title = TeacherProfile.Title.PROF_DR
        name = raw[len("Prof.Dr. "):]
    elif raw.startswith("Dr. "):
        title = TeacherProfile.Title.DR
        name = raw[len("Dr. "):]
    elif raw.startswith("Mr. "):
        title = TeacherProfile.Title.LECTURER
        name = raw[len("Mr. "):]
    elif raw.startswith("Ms. "):
        title = TeacherProfile.Title.LECTURER
        name = raw[len("Ms. "):]
    else:
        # No salutation → no title, visiting faculty
        title = ""
        is_visiting = True
        name = raw

    name = name.strip()
    if not name:
        return None

    parts = name.split()
    if len(parts) == 1:
        first_name, last_name = parts[0], ""
    else:
        first_name, last_name = parts[0], " ".join(parts[1:])

    return title, first_name, last_name, is_visiting


def _teacher_username(first_name, last_name, used):
    """Build a unique username from the teacher's name."""
    import re

    base = re.sub(r"[^a-z0-9]+", "", f"{first_name}{last_name}".lower())
    if not base:
        base = "teacher"
    candidate = base
    n = 2
    while candidate in used or User.objects.filter(username=candidate).exists():
        candidate = f"{base}{n}"
        n += 1
    used.add(candidate)
    return candidate


def seed_teachers_from_file(school):
    """Seed teacher accounts + profiles from data/teachers.txt."""
    if not TEACHERS_TXT.is_file():
        raise FileNotFoundError(
            f"Missing teachers roster at {TEACHERS_TXT}. "
            "Place teachers.txt under data/."
        )

    # Remove prior teacher accounts so re-seeds don't leave orphans
    User.objects.filter(role=User.RoleChoices.TEACHER).delete()

    lines = TEACHERS_TXT.read_text(encoding="utf-8").splitlines()
    used_usernames = set()
    profiles = []
    created = 0

    for line in lines:
        parsed = _parse_teacher_line(line)
        if parsed is None:
            continue
        title, first_name, last_name, is_visiting = parsed
        username = _teacher_username(first_name, last_name, used_usernames)
        email_local = username
        email = f"{email_local}@gmail.com"

        user, created_user = User.objects.get_or_create(
            username=username,
            defaults={
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "role": User.RoleChoices.TEACHER,
                "school": school,
            },
        )
        user.email = email
        user.first_name = first_name
        user.last_name = last_name
        user.role = User.RoleChoices.TEACHER
        user.school = school
        user.set_password("teacherpass")
        user.save()

        # Drop prior profile if re-seeding left an orphan link
        TeacherProfile.objects.filter(user=user).delete()
        profile = TeacherProfile.objects.create(
            user=user,
            employee_id=TeacherProfile.generate_employee_id(),
            title=title,
            is_visiting=is_visiting,
            is_active=True,
        )
        profiles.append(profile)
        created += 1
        created_flag = "new" if created_user else "updated"
        print(f"  {profile.employee_id} {profile.ranked_name} ({created_flag}, visiting={is_visiting})")

    print(f"Teachers seeded from file: {created} created.")
    return profiles


def seed_class_sessions_from_semesters(session, courses_by_code, teachers):
    """
    Create ClassSessions for Fall (odd absolute semesters 1/3/5/7) from
    data/semesters.xlsx. Skips rows whose Subject is not already seeded.
    """
    if not SEMESTERS_XLSX.is_file():
        raise FileNotFoundError(
            f"Missing semester workbook at {SEMESTERS_XLSX}. "
            "Place semesters.xlsx under data/."
        )
    if not teachers:
        raise ValueError("Need at least one teacher to assign class sessions.")

    wb = load_workbook(SEMESTERS_XLSX, read_only=True, data_only=True)
    try:
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
    finally:
        wb.close()

    subjects_by_code = {s.code: s for s in Subject.objects.all()}
    created = 0
    skipped_even = 0
    skipped_missing_subject = 0
    skipped_unknown_program = 0
    skipped_bad_row = 0
    active_levels = set()  # CourseLevel ids that receive sessions

    for row in rows:
        if not row or not row[0]:
            skipped_bad_row += 1
            continue
        code = str(row[0]).strip()
        if not code or code == "None":
            skipped_bad_row += 1
            continue

        credits = row[2] if len(row) > 2 else None
        program = str(row[4] or "").strip() if len(row) > 4 else ""
        year = row[5] if len(row) > 5 else None
        sem_in_year = row[6] if len(row) > 6 else None
        if year is None or sem_in_year is None or not program:
            skipped_bad_row += 1
            continue

        try:
            abs_sem = _absolute_semester(year, sem_in_year)
        except (TypeError, ValueError):
            skipped_bad_row += 1
            continue

        # Fall 2026 → odd absolute semesters only
        if abs_sem % 2 == 0:
            skipped_even += 1
            continue

        course_code = PROGRAM_TO_COURSE_CODE.get(program)
        if course_code is None:
            skipped_unknown_program += 1
            print(f"  SKIP {code}: unknown program {program!r}")
            continue

        course = courses_by_code.get(course_code)
        if course is None:
            skipped_unknown_program += 1
            print(f"  SKIP {code}: course {course_code} not seeded")
            continue

        subject = subjects_by_code.get(code)
        if subject is None:
            skipped_missing_subject += 1
            continue

        course_level = course.levels.get(level=abs_sem)
        active_levels.add(course_level.id)
        ClassSession.objects.create(
            session=session,
            subject=subject,
            teacher=random.choice(teachers),
            course_level=course_level,
            periods_per_week=_periods_per_week_from_credits(credits),
        )
        created += 1

    # Mark odd levels that have sessions with a default cohort size
    CourseLevel.objects.filter(id__in=active_levels).update(student_count=60)

    # Mark each scheduled cohort as a running semester (default Day shift)
    offering_pairs = (
        ClassSession.objects.filter(session=session)
        .values_list("course_level_id", flat=True)
        .distinct()
    )
    offerings_created = 0
    for course_level_id in offering_pairs:
        _, was_created = CourseLevelOffering.objects.get_or_create(
            session=session,
            course_level_id=course_level_id,
            defaults={"shift": CourseLevelOffering.Shift.DAY},
        )
        if was_created:
            offerings_created += 1

    print(
        f"ClassSessions seeded: {created} created "
        f"(skipped even={skipped_even}, missing subject={skipped_missing_subject}, "
        f"unknown program={skipped_unknown_program}, bad row={skipped_bad_row}); "
        f"running offerings: {offerings_created} created."
    )
    return created


def run_tests_and_seed():
    print("Running Tests and Seeding DB...")
    
    # Clean up previous seed if running multiple times
    from timetable.models import TimetableSlot, Timetable, DraftChangeSet, DraftMove

    DraftMove.objects.all().delete()
    DraftChangeSet.objects.all().delete()
    TimetableSlot.objects.all().delete()
    Timetable.objects.all().delete()
    Room.objects.all().delete()
    ClassSession.objects.all().delete()
    Subject.objects.all().delete()
    Course.objects.all().delete()
    TeacherAvailability.objects.all().delete()
    TeacherProfile.objects.all().delete()
    Department.objects.all().delete()
    Session.objects.all().delete()

    school, _ = School.objects.get_or_create(
        code='default',
        defaults={'name': 'Default School', 'is_active': True},
    )
    
    # 1. Create the only academic session
    print("Seeding Sessions...")
    s1 = Session.objects.create(
        name="Fall 2026",
        start_date=datetime.date(2026, 8, 1), end_date=datetime.date(2026, 12, 15),
        is_active=True, school=school,
    )

    # Test Session uniqueness constraint on is_active (scoped per school); not saved
    s_dup = Session(
        name="Duplicate Active",
        start_date=datetime.date(2026, 8, 1), end_date=datetime.date(2026, 12, 15),
        is_active=True, school=school,
    )
    try:
        s_dup.clean()
        print("FAIL: Expected ValidationError for second active session")
    except ValidationError as e:
        print(f"PASS: Validation error raised for second active session: {e}")

    # 2. Create Departments
    print("Seeding Departments...")
    department_specs = [
        ("Department of Architecture", "ARCH"),
        ("Department of Artificial Intelligence", "AI"),
        ("Department of Chemical Science and Engineering", "CHE"),
        ("Department of Civil Engineering", "CE"),
        ("Department of Computer Science and Engineering", "CSE"),
        ("Department of Electrical and Electronics Engineering", "EEE"),
        ("Department of Environmental Engineering", "ENV"),
        ("Department of Geomatics Engineering", "GE"),
        ("Department of Health Informatics", "HI"),
        ("Department of Mechanical Engineering", "ME"),
    ]
    departments = {
        code: Department.objects.create(
            name=name,
            code=code,
            description=name,
            school=school,
        )
        for name, code in department_specs
    }
    arch = departments["ARCH"]
    ai = departments["AI"]
    che = departments["CHE"]
    ce = departments["CE"]
    cse = departments["CSE"]
    eee = departments["EEE"]
    env = departments["ENV"]
    ge = departments["GE"]
    hi = departments["HI"]
    me = departments["ME"]

    # 3. Create Rooms
    print("Seeding Rooms...")
    lecture = Room.RoomType.LECTURE
    lab = Room.RoomType.LAB
    seminar = Room.RoomType.SEMINAR
    computer_lab = Room.RoomType.COMPUTER_LAB

    # name, code, building, floor, capacity, room_type, department
    # Block 9: floors 3/4 → CSE, floor 2 → Geomatics; LAB_305 → CSE (block 9).
    room_specs = [
        ("9-302", "9-302", "9", "3", 60, computer_lab, cse),
        ("9-304", "9-304", "9", "3", 60, lecture, cse),
        ("9-310", "9-310", "9", "3", 60, lecture, cse),
        ("9-402", "9-402", "9", "4", 60, lecture, cse),
        ("9-403", "9-403", "9", "4", 60, lecture, cse),
        ("9-404", "9-404", "9", "4", 60, lecture, cse),
        ("LAB_305", "LAB_305", "9", "3", 30, lab, cse),
        ("9-301", "9-301", "9", "3", 60, seminar, cse),  # Graduate Room
        ("9-202", "9-202", "9", "2", 60, lecture, ge),
        ("9-203", "9-203", "9", "2", 60, lecture, ge),
        ("9-203A", "9-203A", "9", "2", 60, lecture, ge),
        ("9-201", "9-201", "9", "2", 60, lab, ge),  # Simulation Lab
        ("9-Active_Learning_LAB", "9-AL-LAB", "9", "", 30, computer_lab, cse),
        ("10-103", "10-103", "10", "1", 60, lecture, arch),
        ("10-106", "10-106", "10", "1", 60, lecture, arch),
        ("10-202", "10-202", "10", "2", 30, lecture, arch),
        ("10-102", "10-102", "10", "1", 60, lecture, arch),
        ("10-201", "10-201", "10", "2", 60, lecture, arch),
        ("10-107", "10-107", "10", "1", 60, lecture, arch),
        ("Archi Block (Shed)", "ARCHI-SHED", "Archi Block", "", 60, lab, arch),
        ("6-208", "6-208", "6", "2", 60, lecture, eee),
        ("6-202", "6-202", "6", "2", 60, lecture, eee),
        ("6-203", "6-203", "6", "2", 60, lecture, eee),
        ("6-209", "6-209", "6", "2", 60, lecture, eee),
        ("6-S3", "6-S3", "6", "", 60, seminar, eee),
        ("6-S4", "6-S4", "6", "", 30, seminar, eee),
        ("6-S5", "6-S5", "6", "", 30, seminar, eee),
        ("6-S6", "6-S6", "6", "", 60, seminar, eee),
        ("Electrical Lab", "ELEC-LAB", "", "", 60, lab, None),
        ("8-505", "8-505", "8", "5", 60, lecture, me),
        ("8-204", "8-204", "8", "2", 60, lecture, me),
        ("8-502", "8-502", "8", "5", 30, lecture, me),
        ("8-503", "8-503", "8", "5", 30, lecture, me),
        ("11-104", "11-104", "11", "1", 60, lecture, ce),
        ("11-110", "11-110", "11", "1", 60, lecture, ce),
        ("11-105", "11-105", "11", "1", 60, lecture, ce),
        ("3-LUPIC Lab", "3-LUPIC-L", "3", "", 30, lab, None),
        ("3-LUPIC Class Room", "3-LUPIC-C", "3", "", 30, lecture, None),
        ("TTC", "TTC", "", "", 60, lecture, None),
        ("Drawing Hall", "DRAW-HALL", "", "", 60, lecture, None),
        ("Workshop", "WORKSHOP", "", "", 60, lab, None),
        ("Rinpoche-1", "RINPOCHE-1", "", "", 60, lecture, None),
        ("Rinpoche-2", "RINPOCHE-2", "", "", 60, lecture, None),
    ]
    for name, code, building, floor, capacity, room_type, dept in room_specs:
        Room.objects.create(
            name=name,
            code=code,
            building=building,
            floor=floor,
            capacity=capacity,
            room_type=room_type,
            department=dept,
            school=school,
        )

    # Test SET_NULL on Room when Department is deleted
    test_dept = Department.objects.create(name="Test Dept", code="TEST", school=school)
    test_room = Room.objects.create(name="Test Room", code="TEST-RM", capacity=50, department=test_dept, school=school)
    test_dept.delete()
    test_room.refresh_from_db()
    if test_room.department is None:
        print("PASS: Room department set to NULL upon department deletion.")
    else:
        print("FAIL: Expected Room department to be NULL.")
    test_room.delete()

    # 4. Create Subjects from KU SOE scraped workbook
    print("Seeding Subjects...")
    seed_subjects_from_excel(departments)

    # 5. Create Courses (bachelor's programs per department); .save() auto-creates levels 1–8
    print("Seeding Courses...")
    course_specs = [
        # Computer Science and Engineering
        ("BE in Computer Engineering", "BE-CE", cse),
        ("Bachelor of Information Technology (BIT)", "BIT", cse),
        ("Bachelor of Information Technology (BIT) – Double Degree", "BIT-DD", cse),
        ("BSc in Computer Science", "BSC-CS", cse),
        ("B.Tech in Cybersecurity", "BTECH-CYBER", cse),
        # Electrical and Electronics Engineering
        ("BE in Electrical and Electronics Engineering", "BE-EEE", eee),
        ("BE in Communication Engineering", "BE-COMM", eee),
        # Mechanical Engineering (tracks)
        ("BE in Mechanical Engineering (Automobile)", "BE-ME-AUTO", me),
        ("BE in Mechanical Engineering (Design & Manufacturing)", "BE-ME-DM", me),
        ("BE in Mechanical Engineering (Energy Technology)", "BE-ME-ET", me),
        ("BE in Mechanical Engineering (Hydropower)", "BE-ME-HP", me),
        # Geomatics Engineering
        ("BE in Geomatics Engineering", "BE-GE", ge),
        # Architecture
        ("Bachelor in Heritage Conservation (BHC)", "BHC", arch),
        ("Bachelor of Architecture (B.Arch)", "BARCH", arch),
        # Chemical Science and Engineering
        ("BE in Chemical Engineering", "BE-CHE", che),
        # Civil Engineering
        ("BE in Civil Engineering", "BE-CIV", ce),
        ("BE in Mining Engineering", "BE-MIN", ce),
        # Artificial Intelligence
        ("Bachelor of Technology (B.Tech) in Artificial Intelligence", "BTECH-AI", ai),
        # Environmental Engineering
        ("BE in Environmental Engineering", "BE-ENV", env),
        # Health Informatics: no bachelor's program listed
    ]
    courses_by_code = {}
    for program_name, code, dept in course_specs:
        course = Course.objects.create(
            name=program_name,
            code=code,
            department=dept,
            is_active=True,
        )
        courses_by_code[code] = course

    # 6. Create TeacherProfiles from KU roster
    print("Seeding Teacher Profiles...")
    teacher_profiles = seed_teachers_from_file(school)

    # 7. ClassSessions for odd absolute semesters from semesters.xlsx
    print("Seeding ClassSessions...")
    random.seed(42)  # stable teacher assignment across re-seeds
    cs_count = seed_class_sessions_from_semesters(s1, courses_by_code, teacher_profiles)

    print(
        f"Seed complete! Subjects: {Subject.objects.count()}, "
        f"Courses: {Course.objects.count()}, Teachers: {TeacherProfile.objects.count()}, "
        f"ClassSessions: {cs_count}, Sessions: {Session.objects.count()}."
    )

if __name__ == '__main__':
    run_tests_and_seed()
