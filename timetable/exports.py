"""
timetable/exports.py

In-memory PDF and Excel generation for timetable grids.

Layout matches the on-screen academic grid:
  - Time slots across the top
  - Days down the left
  - Colour-coded cells with subject code, name, teacher, room, and time
  - Lunch break column when consecutive periods have a wall-clock gap
"""

from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# reportlab is already locked in requirements.txt for timetable PDFs because it
# is pure Python and avoids the system Cairo/Pango dependencies of weasyprint.
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


DAY_NAMES = {
    1: "Monday",
    2: "Tuesday",
    3: "Wednesday",
    4: "Thursday",
    5: "Friday",
}


# Subject colour palette — mirrors static/css/timetable_grid.css
# Each entry is (background_hex, text_hex).
SUBJECT_PALETTE = [
    ("E8F0FE", "1A3E72"),
    ("FCE8E6", "7A1B14"),
    ("E6F4EA", "145225"),
    ("FEF7E0", "7A5900"),
    ("F3E8FD", "4A1B72"),
    ("E0F7FA", "004D56"),
    ("FCE4EC", "6E0D2E"),
    ("E8EAF6", "1A237E"),
    ("FFF3E0", "7A4100"),
    ("E0F2F1", "004D40"),
]


def _subject_colour(subject_code):
    """Return (background_hex, text_hex) for a subject code (deterministic)."""
    index = sum(ord(c) for c in (subject_code or "")) % len(SUBJECT_PALETTE)
    return SUBJECT_PALETTE[index]


def _cell_colour(slots):
    """Colour for a grid cell, based on the first slot's subject."""
    if not slots:
        return None
    return _subject_colour(slots[0].class_session.subject.code)


def _grid_data(slots, timeslots):
    """Build transposed grid data: days as rows, periods as columns."""
    days = sorted({ts.day_of_week for ts in timeslots})
    periods = sorted({ts.period_number for ts in timeslots})
    period_times = {}
    for ts in timeslots:
        period_times.setdefault(ts.period_number, (ts.start_time, ts.end_time))

    period_meta = []
    for index, period in enumerate(periods):
        start_time, end_time = period_times[period]
        meta = {
            "number": period,
            "start_time": start_time,
            "end_time": end_time,
            "break_after": None,
        }
        if index < len(periods) - 1:
            next_start, _ = period_times[periods[index + 1]]
            if end_time != next_start:
                meta["break_after"] = (end_time, next_start)
        period_meta.append(meta)

    grid = defaultdict(list)
    for slot in slots:
        grid[(slot.timeslot.day_of_week, slot.timeslot.period_number)].append(slot)

    return days, period_meta, grid


def _escape(text):
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _teacher_name(slot):
    if not slot.teacher:
        return ""
    return (
        getattr(slot.teacher, "display_name", None)
        or slot.teacher.user.get_full_name()
        or slot.teacher.user.username
    )


def _slot_lines(slot, scope, period_meta=None):
    """Lines matching the on-screen activity card."""
    subject = slot.class_session.subject
    lines = [
        subject.code,
        subject.name,
    ]
    teacher = _teacher_name(slot)
    if teacher:
        lines.append(teacher)
    lines.append(slot.room.name)
    if period_meta is not None:
        lines.append(
            f"{period_meta['start_time']:%H:%M}–{period_meta['end_time']:%H:%M}"
        )
    if scope != "section":
        lines.append(slot.class_session.section.name)
    if slot.is_locked:
        lines.append("Locked")
    return lines


def _cell_text(slots, scope, period_meta=None):
    if not slots:
        return ""
    return "\n\n".join(
        "\n".join(_slot_lines(slot, scope, period_meta)) for slot in slots
    )


def _pdf_cell_paragraph(slots, scope, period_meta, styles):
    """Rich PDF cell: bold subject code, then name / teacher / room / time."""
    if not slots:
        return Paragraph("—", styles["empty"])

    blocks = []
    for slot in slots:
        subject = slot.class_session.subject
        parts = [
            f"<b>{_escape(subject.code)}</b>",
            _escape(subject.name),
        ]
        teacher = _teacher_name(slot)
        if teacher:
            parts.append(_escape(teacher))
        parts.append(_escape(slot.room.name))
        parts.append(
            _escape(
                f"{period_meta['start_time']:%H:%M}–{period_meta['end_time']:%H:%M}"
            )
        )
        if scope != "section":
            parts.append(_escape(slot.class_session.section.name))
        if slot.is_locked:
            parts.append("Locked")
        blocks.append("<br/>".join(parts))
    return Paragraph("<br/><br/>".join(blocks), styles["cell"])


def _column_plan(period_meta):
    """Flat list of columns: period entries and optional lunch columns."""
    columns = []
    for meta in period_meta:
        columns.append({"kind": "period", "meta": meta})
        if meta["break_after"]:
            columns.append({
                "kind": "lunch",
                "meta": {
                    "start_time": meta["break_after"][0],
                    "end_time": meta["break_after"][1],
                },
            })
    return columns


def export_timetable_pdf(*, slots, timeslots, title, subtitle, scope):
    """Return PDF bytes for a timetable grid matching the on-screen layout."""
    days, period_meta, grid = _grid_data(slots, timeslots)
    columns = _column_plan(period_meta)

    buffer = BytesIO()
    page_width, page_height = landscape(letter)
    margin = 0.4 * inch
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=margin,
        leftMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
    )

    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "TFTitle",
            parent=base["Title"],
            fontName="Helvetica-Bold",
            fontSize=16,
            textColor=colors.HexColor("#111827"),
            alignment=TA_CENTER,
            spaceAfter=4,
        ),
        "subtitle": ParagraphStyle(
            "TFSubtitle",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=9,
            textColor=colors.HexColor("#6B7280"),
            alignment=TA_CENTER,
            spaceAfter=10,
        ),
        "header": ParagraphStyle(
            "TFHeader",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7,
            textColor=colors.HexColor("#111827"),
            alignment=TA_CENTER,
            leading=9,
        ),
        "day": ParagraphStyle(
            "TFDay",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            textColor=colors.HexColor("#111827"),
            alignment=TA_LEFT,
            leading=10,
        ),
        "cell": ParagraphStyle(
            "TFCell",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=6.5,
            leading=8,
            alignment=TA_LEFT,
        ),
        "empty": ParagraphStyle(
            "TFEmpty",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=7,
            textColor=colors.HexColor("#9CA3AF"),
            alignment=TA_CENTER,
        ),
        "lunch": ParagraphStyle(
            "TFLunch",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7,
            textColor=colors.HexColor("#7A5900"),
            alignment=TA_CENTER,
            leading=9,
        ),
    }

    story = [
        Paragraph(_escape(title), styles["title"]),
        Paragraph(_escape(subtitle), styles["subtitle"]),
        Spacer(1, 0.05 * inch),
    ]

    # Header row: Day/Time + each period (and lunch) column
    header = [Paragraph("Day / Time", styles["header"])]
    for column in columns:
        if column["kind"] == "lunch":
            lunch = column["meta"]
            lunch_label = f"{lunch['start_time']:%H:%M}–{lunch['end_time']:%H:%M}"
            header.append(
                Paragraph(f"Lunch<br/>{_escape(lunch_label)}", styles["lunch"])
            )
        else:
            meta = column["meta"]
            time_label = f"{meta['start_time']:%H:%M}–{meta['end_time']:%H:%M}"
            header.append(
                Paragraph(
                    f"{_escape(time_label)}"
                    f"<br/><font size='6' color='#9CA3AF'>Period {meta['number']}</font>",
                    styles["header"],
                )
            )

    table_data = [header]
    cell_styles = []

    for row_index, day in enumerate(days, start=1):
        row = [Paragraph(DAY_NAMES.get(day, f"Day {day}"), styles["day"])]
        col_index = 1
        for column in columns:
            if column["kind"] == "lunch":
                lunch = column["meta"]
                lunch_label = (
                    f"{lunch['start_time']:%H:%M}–{lunch['end_time']:%H:%M}"
                )
                if row_index == 1:
                    row.append(
                        Paragraph(
                            f"Lunch<br/>{_escape(lunch_label)}",
                            styles["lunch"],
                        )
                    )
                    if len(days) > 1:
                        cell_styles.append(
                            ("SPAN", (col_index, 1), (col_index, len(days)))
                        )
                    cell_styles.append(
                        (
                            "BACKGROUND",
                            (col_index, 1),
                            (col_index, len(days)),
                            colors.HexColor("#F4F1E7"),
                        )
                    )
                else:
                    row.append("")
                col_index += 1
                continue

            meta = column["meta"]
            cell_slots = grid[(day, meta["number"])]
            row.append(_pdf_cell_paragraph(cell_slots, scope, meta, styles))
            colour = _cell_colour(cell_slots)
            if colour:
                bg, text = colour
                cell_styles.append(
                    (
                        "BACKGROUND",
                        (col_index, row_index),
                        (col_index, row_index),
                        colors.HexColor(f"#{bg}"),
                    )
                )
                cell_styles.append(
                    (
                        "TEXTCOLOR",
                        (col_index, row_index),
                        (col_index, row_index),
                        colors.HexColor(f"#{text}"),
                    )
                )
            col_index += 1
        table_data.append(row)

    available_width = page_width - (2 * margin)
    day_width = 0.95 * inch
    lunch_width = 0.55 * inch
    period_count = sum(1 for c in columns if c["kind"] == "period")
    lunch_count = sum(1 for c in columns if c["kind"] == "lunch")
    remaining = available_width - day_width - (lunch_count * lunch_width)
    period_width = remaining / max(period_count, 1)

    col_widths = [day_width]
    for column in columns:
        col_widths.append(lunch_width if column["kind"] == "lunch" else period_width)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F9FAFB")),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F9FAFB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (1, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
    ] + cell_styles))
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def export_timetable_xlsx(*, slots, timeslots, title, subtitle, scope):
    """Return XLSX bytes matching the on-screen transposed timetable grid."""
    days, period_meta, grid = _grid_data(slots, timeslots)
    columns = _column_plan(period_meta)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Timetable"

    max_col = 1 + len(columns)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    sheet.cell(1, 1, title)
    sheet.cell(2, 1, subtitle)
    sheet.cell(1, 1).font = Font(size=16, bold=True, color="111827")
    sheet.cell(2, 1).font = Font(size=10, color="6B7280")
    sheet.cell(1, 1).alignment = Alignment(horizontal="center")
    sheet.cell(2, 1).alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="F9FAFB")
    day_fill = PatternFill("solid", fgColor="F9FAFB")
    lunch_fill = PatternFill("solid", fgColor="F4F1E7")
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_row = 4
    sheet.cell(header_row, 1, "Day / Time")
    for index, column in enumerate(columns, start=2):
        if column["kind"] == "lunch":
            lunch = column["meta"]
            value = f"Lunch\n{lunch['start_time']:%H:%M}–{lunch['end_time']:%H:%M}"
            cell = sheet.cell(header_row, index, value)
            cell.fill = lunch_fill
            cell.font = Font(bold=True, color="7A5900", size=9)
        else:
            meta = column["meta"]
            value = (
                f"{meta['start_time']:%H:%M}–{meta['end_time']:%H:%M}\n"
                f"Period {meta['number']}"
            )
            cell = sheet.cell(header_row, index, value)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="111827", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    day_header = sheet.cell(header_row, 1)
    day_header.fill = header_fill
    day_header.font = Font(bold=True, color="111827", size=9)
    day_header.alignment = Alignment(horizontal="center", vertical="center")
    day_header.border = border

    for row_offset, day in enumerate(days):
        row_index = header_row + 1 + row_offset
        day_cell = sheet.cell(row_index, 1, DAY_NAMES.get(day, f"Day {day}"))
        day_cell.fill = day_fill
        day_cell.font = Font(bold=True, color="111827", size=10)
        day_cell.alignment = Alignment(horizontal="left", vertical="center")
        day_cell.border = border

        for col_index, column in enumerate(columns, start=2):
            if column["kind"] == "lunch":
                if row_offset == 0:
                    lunch = column["meta"]
                    cell = sheet.cell(
                        row_index,
                        col_index,
                        f"Lunch\n{lunch['start_time']:%H:%M}–{lunch['end_time']:%H:%M}",
                    )
                    cell.fill = lunch_fill
                    cell.font = Font(bold=True, color="7A5900", size=9)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True, textRotation=90
                    )
                    cell.border = border
                    if len(days) > 1:
                        sheet.merge_cells(
                            start_row=row_index,
                            start_column=col_index,
                            end_row=header_row + len(days),
                            end_column=col_index,
                        )
                continue

            meta = column["meta"]
            cell_slots = grid[(day, meta["number"])]
            cell = sheet.cell(
                row_index,
                col_index,
                _cell_text(cell_slots, scope, meta),
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            colour = _cell_colour(cell_slots)
            if colour:
                bg, text = colour
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(color=text, size=9)
            else:
                cell.font = Font(color="9CA3AF", size=9)

    sheet.freeze_panes = "B5"
    sheet.column_dimensions["A"].width = 14
    for col_index, column in enumerate(columns, start=2):
        letter = get_column_letter(col_index)
        sheet.column_dimensions[letter].width = 8 if column["kind"] == "lunch" else 22
    for row_index in range(header_row + 1, header_row + len(days) + 1):
        sheet.row_dimensions[row_index].height = 78
    sheet.row_dimensions[header_row].height = 36

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
